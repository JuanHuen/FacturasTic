"""
app.py — Ingreso de Facturas y Boletas (Web)
=============================================
Streamlit + Supabase

Requisitos:
    pip install streamlit supabase requests pandas

Correr local:
    streamlit run app.py
"""

import streamlit as st
import requests
import re
from datetime import datetime
from typing import Optional, List, Dict

# ─── CONFIGURACIÓN SUPABASE ───────────────────────────────────────────────────
# Estas variables las pones en Streamlit Cloud > Settings > Secrets así:
# [supabase]
# url = "https://xxxx.supabase.co"
# key = "eyJ..."

def get_supabase_config():
    try:
        url = st.secrets["supabase"]["url"]
        key = st.secrets["supabase"]["key"]
    except Exception:
        url = ""
        key = ""
    return url, key

SUPABASE_URL, SUPABASE_KEY = get_supabase_config()

MESES = ["Enero","Febrero","Marzo","Abril","Mayo","Junio",
         "Julio","Agosto","Septiembre","Octubre","Noviembre","Diciembre"]
DEFAULT_GRUPOCORP_OPTIONS = ["APLICACIONES","INFRAESTRUCTURA","TELECOM","SOFTWARE","CIBERSEGURIDAD"]
DEFAULT_OPER_OPTIONS      = ["Operativo","Iniciativa"]

# ─── SUPABASE HELPERS ────────────────────────────────────────────────────────
def _headers():
    return {
        "apikey": SUPABASE_KEY,
        "Authorization": f"Bearer {SUPABASE_KEY}",
        "Content-Type": "application/json",
        "Prefer": "return=representation",
    }

def sb_select(table: str, query: str = "") -> list:
    url = f"{SUPABASE_URL}/rest/v1/{table}?{query}&limit=5000"
    try:
        r = requests.get(url, headers=_headers(), timeout=10)
        r.raise_for_status()
        return r.json()
    except Exception as e:
        st.error(f"Error leyendo {table}: {e}")
        return []

def sb_insert(table: str, data: dict) -> bool:
    url = f"{SUPABASE_URL}/rest/v1/{table}"
    try:
        r = requests.post(url, headers=_headers(), json=data, timeout=10)
        r.raise_for_status()
        return True
    except Exception as e:
        st.error(f"Error guardando en {table}: {e}")
        return False

def sb_upsert(table: str, data: dict, on_conflict: str = "") -> bool:
    url = f"{SUPABASE_URL}/rest/v1/{table}"
    headers = _headers()
    if on_conflict:
        headers["Prefer"] = f"resolution=merge-duplicates,return=representation"
        url += f"?on_conflict={on_conflict}"
    try:
        r = requests.post(url, headers=headers, json=data, timeout=10)
        r.raise_for_status()
        return True
    except Exception as e:
        st.error(f"Error en upsert {table}: {e}")
        return False

# ─── HELPERS DE LÓGICA (igual que MyApp.py) ──────────────────────────────────
def limpiar_ruc(s) -> str:
    if s is None: return ""
    st_val = str(s).strip()
    if st_val.endswith(".0"): st_val = st_val[:-2]
    return "".join(ch for ch in st_val if ch.isdigit())

def _canon_key(s: Optional[str]) -> str:
    if s is None: return ""
    return re.sub(r'\s+', ' ', str(s).replace('\u00A0', ' ')).strip()

def to_list(v) -> List[str]:
    if v is None: return []
    s = str(v).strip()
    if s == "" or s.lower() == "nan": return []
    parts = [p.strip() for p in (s.split(';') if ';' in s else s.split(','))]
    seen = set(); out = []
    for p in parts:
        if p and p.lower() != "nan" and p not in seen:
            seen.add(p); out.append(p)
    return out

def get_exchange_rate() -> float:
    try:
        r = requests.get('https://api.exchangerate-api.com/v4/latest/USD', timeout=6)
        r.raise_for_status()
        rate = r.json().get('rates', {}).get('PEN')
        if rate: return round(float(rate), 2)
    except Exception:
        pass
    return 3.40

# ─── CARGAR MAESTRO DESDE SUPABASE ───────────────────────────────────────────
@st.cache_data(ttl=300)
def cargar_maestro() -> Dict:
    """Carga el Maestro desde Supabase y construye el LOOKUP."""
    rows = sb_select("maestro")
    if not rows:
        return {}
    lookup: Dict = {}
    for row in rows:
        ruc = limpiar_ruc(row.get("ruc", ""))
        if not ruc: continue
        provs = to_list(row.get("proveedor", ""))
        if not provs: continue
        grupos = to_list(row.get("grupodesccorporativa", "")) or [""]
        descs  = to_list(row.get("descripciones", ""))
        facts  = to_list(row.get("facturas", ""))
        grpcorp = to_list(row.get("grupocorpclasitiposervtic", ""))
        oper    = to_list(row.get("operativoiniciativa", ""))
        moneda  = to_list(row.get("moneda", ""))
        tipo    = to_list(row.get("tipo", ""))

        if ruc not in lookup: lookup[ruc] = {}
        for prov in provs:
            pk = _canon_key(prov)
            if pk not in lookup[ruc]: lookup[ruc][pk] = {}
            for g in grupos:
                gk = _canon_key(g)
                if gk not in lookup[ruc][pk]:
                    lookup[ruc][pk][gk] = {"descripciones":[],"facturas":[],"grpcorp":[],"oper":[],"moneda":[],"tipo":[]}
                cell = lookup[ruc][pk][gk]
                for d in descs:
                    if d and d not in cell["descripciones"]: cell["descripciones"].append(d)
                for f in facts:
                    if f and f not in cell["facturas"]: cell["facturas"].append(f)
                for v in grpcorp:
                    if v and v not in cell["grpcorp"]: cell["grpcorp"].append(v)
                for v in oper:
                    if v and v not in cell["oper"]: cell["oper"].append(v)
                for v in moneda:
                    if v and v not in cell["moneda"]: cell["moneda"].append(v)
                for v in tipo:
                    if v and v not in cell["tipo"]: cell["tipo"].append(v)
    return lookup

def get_rucs(lookup: Dict) -> List[str]:
    return sorted(lookup.keys())

def get_proveedores(lookup: Dict, ruc: str) -> List[str]:
    if not ruc or ruc not in lookup: return []
    return [p for p in lookup[ruc].keys() if p]

def get_grupos(lookup: Dict, ruc: str, prov: str) -> List[str]:
    if not ruc or not prov: return []
    prov_data = lookup.get(ruc, {}).get(_canon_key(prov), {})
    return [g if g else "<Sin grupo>" for g in prov_data.keys()]

def get_cell_data(lookup: Dict, ruc: str, prov: str, grp: str) -> dict:
    grp_real = "" if grp == "<Sin grupo>" else grp
    try:
        return lookup[ruc][_canon_key(prov)][_canon_key(grp_real)]
    except (KeyError, TypeError):
        return {}

# ─── PÁGINA INGRESAR FACTURA ─────────────────────────────────────────────────
def pagina_ingresar(lookup: Dict):
    st.header("📄 Ingreso de Facturas")

    # Contador para resetear widgets (cambia la key de cada widget al guardar)
    form_n = st.session_state.get("form_n", 0)

    rucs = get_rucs(lookup)
    anio_actual = datetime.now().year

    col1, col2 = st.columns(2)

    # ── Fechas
    with col1:
        st.subheader("Fechas")
        fecha_em = st.date_input("Fecha de Emisión",    value=datetime.now())
        fecha_b  = st.date_input("Fecha de Trámite IT", value=datetime.now())
        fecha_c  = st.date_input("Fecha de Imputación IT", value=datetime.now())

    # ── RUC / Proveedor / Grupo  (keys con form_n para resetear al guardar)
    fn = form_n
    with col2:
        st.subheader("Identificación")
        ruc_input = st.selectbox("RUC", options=[""] + rucs, index=0, key=f"ruc_{fn}")
        ruc = limpiar_ruc(ruc_input)

        proveedores = get_proveedores(lookup, ruc)
        prov = st.selectbox("Proveedor", options=proveedores if proveedores else [""], key=f"prov_{fn}")

        grupos = get_grupos(lookup, ruc, prov)
        grp_disp = st.selectbox("GrupoDescCorporativa", options=grupos if grupos else [""], key=f"grp_{fn}")

    # ── Datos de la celda
    cell = get_cell_data(lookup, ruc, prov, grp_disp)
    descs   = cell.get("descripciones", [])
    facts   = cell.get("facturas", [])
    grpcorp_opts = cell.get("grpcorp", DEFAULT_GRUPOCORP_OPTIONS) or DEFAULT_GRUPOCORP_OPTIONS
    oper_opts    = cell.get("oper",    DEFAULT_OPER_OPTIONS)      or DEFAULT_OPER_OPTIONS
    moneda_default = (cell.get("moneda") or ["S"])[0]
    tipo_default   = (cell.get("tipo")   or ["G"])[0]

    st.divider()
    st.subheader("Clasificación y Tipo")
    col3, col4 = st.columns(2)
    with col3:
        grpcorp = st.selectbox("Descripción corta", options=grpcorp_opts, key=f"grpcorp_{fn}")
    with col4:
        oper = st.selectbox("Operativo/Iniciativa", options=oper_opts, key=f"oper_{fn}")

    st.divider()
    st.subheader("Detalles de Factura")

    desc_sel = st.selectbox("Descripción larga", options=descs if descs else [""], key=f"desc_sel_{fn}")
    # Sincronizar texto editable con el selector (solo si el usuario no lo ha editado manualmente)
    desc_key = f"desc_manual_{fn}"
    if st.session_state.get(f"desc_sel_{fn}") != st.session_state.get(f"_prev_desc_sel_{fn}"):
        st.session_state[desc_key] = desc_sel
    st.session_state[f"_prev_desc_sel_{fn}"] = st.session_state.get(f"desc_sel_{fn}")
    desc_manual = st.text_input("O edita la descripción:", key=desc_key).upper()
    descripcion = desc_manual if desc_manual else desc_sel.upper()

    # Contrato
    es_contrato = st.radio("¿Es contrato?", ["No", "Sí"], horizontal=True, key=f"contrato_{fn}") == "Sí"
    plazo = ""
    if es_contrato:
        anios_rango = [str(anio_actual + i) for i in range(-3, 4)]
        cp1, cp2, cp3, cp4 = st.columns(4)
        with cp1:
            mes_ini = st.selectbox("Mes inicio", MESES, key=f"mes_ini_{fn}")
        with cp2:
            anio_ini = st.selectbox("Año inicio", anios_rango, index=3, key=f"anio_ini_{fn}")
        mes_ini_idx = MESES.index(mes_ini)
        anio_ini_int = int(anio_ini)
        opts_fin = []
        for i in range(1, 13):
            m = (mes_ini_idx + i) % 12
            a = anio_ini_int + (mes_ini_idx + i) // 12
            opts_fin.append((MESES[m], str(a)))
        mes_fin_default  = mes_ini
        anio_fin_default = str(anio_ini_int + 1)
        opts_mes_fin  = [o[0] for o in opts_fin]
        opts_anio_fin = sorted(set(o[1] for o in opts_fin))
        idx_mes  = opts_mes_fin.index(mes_fin_default)  if mes_fin_default  in opts_mes_fin  else 0
        idx_anio = opts_anio_fin.index(anio_fin_default) if anio_fin_default in opts_anio_fin else len(opts_anio_fin)-1
        with cp3:
            mes_fin  = st.selectbox("Mes fin",  opts_mes_fin,  index=idx_mes,  key=f"mes_fin_{fn}")
        with cp4:
            anio_fin = st.selectbox("Año fin",  opts_anio_fin, index=idx_anio, key=f"anio_fin_{fn}")
        plazo = f"{mes_ini} {anio_ini} - {mes_fin} {anio_fin}"

    fact_sel = st.selectbox("Factura", options=facts if facts else [""], key=f"fact_sel_{fn}")
    # Sincronizar texto editable con el selector (solo si el usuario no lo ha editado manualmente)
    fact_key = f"fact_manual_{fn}"
    if st.session_state.get(f"fact_sel_{fn}") != st.session_state.get(f"_prev_fact_sel_{fn}"):
        st.session_state[fact_key] = fact_sel
    st.session_state[f"_prev_fact_sel_{fn}"] = st.session_state.get(f"fact_sel_{fn}")
    fact_manual = st.text_input("O edita el número de factura:", key=fact_key)
    factura = fact_manual if fact_manual else fact_sel

    col5, col6 = st.columns(2)
    with col5:
        monto_str = st.text_input("Monto sin IGV", key=f"monto_{fn}")
    with col6:
        moneda = st.radio("Moneda", ["Soles (S)", "Dólares (D)"], horizontal=True, key=f"moneda_{fn}")
        moneda_val = "D" if "D" in moneda else "S"

    col7, col8 = st.columns(2)
    with col7:
        tipo_radio = st.radio("Tipo", ["Gasto", "Inversión"], horizontal=True, key=f"tipo_{fn}")
        tipo_val = "G" if tipo_radio == "Gasto" else "I"
    with col8:
        # El TC se carga de la web solo la primera vez (en main).
        # El usuario puede editarlo y el valor queda en session_state["tc"]
        tc = st.number_input(
            "Tipo de cambio",
            value=float(st.session_state.get("tc", 3.40)),
            format="%.2f",
            step=0.01
        )
        st.session_state["tc"] = tc

    # Total calculado
    try:
        monto_float = float(monto_str.replace(",", ".")) if monto_str else 0.0
        total = round(monto_float * tc, 2) if moneda_val == "D" else round(monto_float, 2)
        st.metric("Total en Soles", f"S/ {total:,.2f}")
    except Exception:
        total = 0.0

    st.divider()

    if st.button("💾 Guardar y siguiente", type="primary", use_container_width=True):
        if not ruc:
            st.error("Debes ingresar un RUC.")
            return
        if not monto_str:
            st.error("Debes ingresar el monto.")
            return

        fa = fecha_em
        mes_anno = f"{fa.strftime('%b')}-{str(fa.year)[2:]}"
        trimestre = "I" if fa.month <= 3 else "II" if fa.month <= 6 else "III" if fa.month <= 9 else "IV"
        grp_real = "" if grp_disp == "<Sin grupo>" else grp_disp

        fila = {
            "fecha_emision":          str(fecha_em),
            "fecha_tramite_ivan":     str(fecha_b),
            "fecha_tramite":          str(fecha_c),
            "mes":                    mes_anno,
            "trimestre":              trimestre,
            "grupodesccorporativa":   grp_real,
            "grupocorpclasitiposervtic": grpcorp,
            "operativoiniciativa":    oper,
            "ruc":                    ruc,
            "proveedor":              prov,
            "descripcion":            descripcion,
            "numero_factura":         factura,
            "tipo":                   "Gasto" if tipo_val == "G" else "Inversión",
            "monto_sin_igv":          monto_float,
            "moneda":                 "DOLARES" if moneda_val == "D" else "SOLES",
            "valor_usd":              tc,
            "monto_total":            total,
            "contrato":               "Sí" if es_contrato else "No",
            "plazo":                  plazo,
        }

        if sb_insert("facturas", fila):
            st.success("✅ Factura guardada correctamente.")
            # Incrementar contador para resetear widgets con keys nuevas (rápido, sin borrar todo)
            st.session_state["form_n"] = st.session_state.get("form_n", 0) + 1
            st.rerun()

# ─── PÁGINA MAESTRO ───────────────────────────────────────────────────────────
def pagina_maestro():
    st.header("📋 Maestro de Proveedores")

    tab1, tab2 = st.tabs(["Ver Maestro", "Agregar / Editar"])

    with tab1:
        rows = sb_select("maestro", "order=proveedor.asc")
        if rows:
            import pandas as pd
            df = pd.DataFrame(rows)
            cols_show = ["ruc","proveedor","grupodesccorporativa","grupocorpclasitiposervtic",
                         "operativoiniciativa","moneda","tipo","descripciones","facturas"]
            cols_show = [c for c in cols_show if c in df.columns]
            st.dataframe(df[cols_show], use_container_width=True, height=400)
        else:
            st.info("No hay datos en el Maestro todavía.")

    with tab2:
        st.subheader("Agregar proveedor al Maestro")
        with st.form("form_maestro"):
            c1, c2 = st.columns(2)
            with c1:
                m_ruc  = st.text_input("RUC")
                m_prov = st.text_input("Proveedor")
                m_grp  = st.text_input("GrupoDescCorporativa")
                m_grpcorp = st.selectbox("GrupoCorpClasiTIpoServTIC", DEFAULT_GRUPOCORP_OPTIONS)
                m_oper    = st.selectbox("OperativoIniciativa", DEFAULT_OPER_OPTIONS)
            with c2:
                m_moneda = st.radio("Moneda", ["S", "D"], horizontal=True)
                m_tipo   = st.radio("Tipo", ["Gasto", "Inversión"], horizontal=True)
                m_descs  = st.text_area("Descripciones (separadas por ;)")
                m_facts  = st.text_area("Facturas (separadas por ;)")

            submitted = st.form_submit_button("Guardar en Maestro", type="primary")
            if submitted:
                if not m_ruc or not m_prov:
                    st.error("RUC y Proveedor son obligatorios.")
                else:
                    data = {
                        "ruc": limpiar_ruc(m_ruc),
                        "proveedor": m_prov,
                        "grupodesccorporativa": m_grp,
                        "grupocorpclasitiposervtic": m_grpcorp,
                        "operativoiniciativa": m_oper,
                        "moneda": m_moneda,
                        "tipo": "Gasto" if m_tipo == "Gasto" else "Inversión",
                        "descripciones": m_descs,
                        "facturas": m_facts,
                    }
                    if sb_insert("maestro", data):
                        st.success("✅ Proveedor agregado.")
                        st.cache_data.clear()
                        st.rerun()

# ─── PÁGINA VER FACTURAS ─────────────────────────────────────────────────────
def pagina_ver_facturas():
    st.header("📊 Facturas Registradas")
    import pandas as pd

    rows = sb_select("facturas", "order=fecha_emision.desc")
    if not rows:
        st.info("No hay facturas registradas todavía.")
        return

    df = pd.DataFrame(rows)

    # Filtros
    col1, col2, col3 = st.columns(3)
    with col1:
        rucs_disp = ["Todos"] + sorted(df["ruc"].dropna().unique().tolist())
        filtro_ruc = st.selectbox("Filtrar por RUC", rucs_disp)
    with col2:
        provs_disp = ["Todos"] + sorted(df["proveedor"].dropna().unique().tolist())
        filtro_prov = st.selectbox("Filtrar por Proveedor", provs_disp)
    with col3:
        anios = ["Todos"] + sorted(df["fecha_emision"].str[:4].dropna().unique().tolist(), reverse=True)
        filtro_anio = st.selectbox("Filtrar por Año", anios)

    if filtro_ruc  != "Todos": df = df[df["ruc"]      == filtro_ruc]
    if filtro_prov != "Todos": df = df[df["proveedor"] == filtro_prov]
    if filtro_anio != "Todos": df = df[df["fecha_emision"].str.startswith(filtro_anio)]

    st.metric("Total facturas", len(df))
    if "monto_total" in df.columns:
        st.metric("Monto total (S/)", f"{df['monto_total'].sum():,.2f}")

    cols_show = ["fecha_emision","proveedor","ruc","descripcion","numero_factura",
                 "tipo","moneda","monto_sin_igv","monto_total","contrato","plazo"]
    cols_show = [c for c in cols_show if c in df.columns]
    st.dataframe(df[cols_show], use_container_width=True, height=500)

    # Exportar Excel con formato
    if st.button("⬇️ Descargar Excel", type="secondary"):
        import io
        from openpyxl import Workbook
        from openpyxl.styles import Font, PatternFill, Alignment, Border, Side

        anio_actual = datetime.now().year
        COLS_EXCEL = [
            ("fecha_emision",           "FECHA DE EMISION"),
            ("fecha_tramite_ivan",       "FECHA DE TRAMITE IVAN"),
            ("fecha_tramite",            "FECHA DE TRAMITE"),
            ("mes",                      "Mes"),
            ("trimestre",                "Trimestre"),
            ("grupodesccorporativa",     "GrupoDescCorporativa"),
            ("grupocorpclasitiposervtic","GrupoCorpClasiTIpoServTIC"),
            ("operativoiniciativa",      "Operativo/Iniciativa"),
            ("ruc",                      "RUC"),
            ("proveedor",                "NOMBRE DEL PROVEEDOR"),
            ("descripcion",              "DESCRIPCION"),
            ("numero_factura",           "NUMERO DE LA FACTURA"),
            ("tipo",                     "Tipo"),
            ("monto_sin_igv",            "MONTO DE LA FACTURA SIN IGV"),
            ("moneda",                   "MONEDA"),
            ("valor_usd",                "Valor USD"),
            ("monto_total",              "MontoTotal"),
        ]
        NUM_COLS = len(COLS_EXCEL)

        wb = Workbook()
        ws = wb.active
        ws.title = f"Facturas {anio_actual}"

        # Fila 1: título
        ws.cell(row=1, column=1).value = f"RELACION DE FACTURAS TRAMITADAS {anio_actual}"
        ws.cell(row=1, column=1).font = Font(bold=True, size=12)
        ws.cell(row=1, column=1).alignment = Alignment(horizontal="center")
        ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=NUM_COLS)

        # Fila 2: encabezados — amarillo con bordes
        fill_yellow = PatternFill("solid", fgColor="FFFF00")
        thin = Side(border_style="thin", color="000000")
        border = Border(left=thin, right=thin, top=thin, bottom=thin)
        for col_idx, (_, header) in enumerate(COLS_EXCEL, start=1):
            c = ws.cell(row=2, column=col_idx)
            c.value = header
            c.font = Font(bold=True)
            c.fill = fill_yellow
            c.border = border
            c.alignment = Alignment(horizontal="center")

        # Filas de datos (fila 3+): sin color, sin bordes
        df_export = df[[col for col, _ in COLS_EXCEL if col in df.columns]]
        for row_idx, row_data in enumerate(df_export.itertuples(index=False), start=3):
            for col_idx, val in enumerate(row_data, start=1):
                ws.cell(row=row_idx, column=col_idx).value = val

        # Autoajustar ancho de columnas (usando índice para evitar problema con celdas mergeadas)
        from openpyxl.utils import get_column_letter
        for col_idx in range(1, NUM_COLS + 1):
            max_len = 0
            for row_idx2 in range(2, ws.max_row + 1):
                cell = ws.cell(row=row_idx2, column=col_idx)
                try:
                    if cell.value:
                        max_len = max(max_len, len(str(cell.value)))
                except Exception:
                    pass
            ws.column_dimensions[get_column_letter(col_idx)].width = min(max_len + 4, 50)

        buf = io.BytesIO()
        wb.save(buf)
        buf.seek(0)
        st.download_button(
            "📥 Descargar archivo Excel",
            buf,
            f"facturas_{anio_actual}.xlsx",
            "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
        )

# ─── MAIN ─────────────────────────────────────────────────────────────────────
def main():
    st.set_page_config(
        page_title="Ingreso de Facturas",
        page_icon="🧾",
        layout="wide",
        initial_sidebar_state="expanded",
    )

    if not SUPABASE_URL or not SUPABASE_KEY:
        st.error("⚠️ Configura las credenciales de Supabase en secrets.toml")
        st.code("""
# .streamlit/secrets.toml
[supabase]
url = "https://xxxx.supabase.co"
key = "eyJ..."
        """)
        st.stop()
#
#
    st.sidebar.title("🧾 Facturas TIC")
    st.sidebar.caption(f"Usuario: {st.secrets.get('usuario', 'Invitado')}")

    pagina = st.sidebar.radio("Navegación", [
        "📄 Ingresar Factura",
        "📊 Ver Facturas",
        "📋 Maestro",
    ])

    # Tipo de cambio al inicio
    if "tc" not in st.session_state:
        with st.spinner("Obteniendo tipo de cambio..."):
            st.session_state["tc"] = get_exchange_rate()

    # Cachear maestro en session para no volver a Supabase en cada rerun del formulario
    if "lookup_cache" not in st.session_state:
        st.session_state["lookup_cache"] = cargar_maestro()
    lookup = st.session_state["lookup_cache"]

    if pagina == "📄 Ingresar Factura":
        pagina_ingresar(lookup)
    elif pagina == "📊 Ver Facturas":
        pagina_ver_facturas()
    elif pagina == "📋 Maestro":
        pagina_maestro()

if __name__ == "__main__":
    main()